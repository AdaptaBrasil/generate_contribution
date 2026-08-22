from __future__ import annotations

import openpyxl
import pandas as pd

from generate_contribution.config import TratamentoConfig
from generate_contribution.treatment import run_tratamento


def _write_input_workbook(path, imeta: pd.DataFrame, idata: pd.DataFrame) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        imeta.to_excel(writer, sheet_name="Metadados", index=False)
        idata.to_excel(writer, sheet_name="Dados", index=False)


def test_run_tratamento_end_to_end(tmp_path, sample_imeta, sample_idata):
    input_path = tmp_path / "input.xlsx"
    _write_input_workbook(input_path, sample_imeta, sample_idata)

    config = TratamentoConfig(
        input=input_path,
        imeta_sheet="Metadados",
        idata_sheet="Dados",
        method_boxcox="forecast",
        sigla="TESTE",
        output_dir=tmp_path / "OUTPUT",
    )

    result = run_tratamento(config)

    outfilex1, outfilex2 = result.output_paths
    assert outfilex1.exists()
    assert outfilex2.exists()

    wb1 = openpyxl.load_workbook(outfilex1)
    assert wb1.sheetnames == ["Descritivo", "Winsorization", "BoxCox"]

    wb2 = openpyxl.load_workbook(outfilex2)
    assert wb2.sheetnames == ["BNivel 7", "Winsorization", "BoxCox", "Normalizado"]

    # BNivel 7 sheet must have GEOCOD, MUN, UF (CLUSTER dropped) + indicator columns
    ws = wb2["BNivel 7"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header[:3] == ["GEOCOD", "MUN", "UF"]
    assert "IND1" in header and "IND2" in header

    assert result.dados_b.shape == (len(sample_idata), 2)
    assert result.data_normal.idata["IND1"].max() <= 1.0
    assert result.data_normal.idata["IND1"].min() >= 0.0
